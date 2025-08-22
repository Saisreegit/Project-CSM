from flask import Flask, request, send_file, jsonify, render_template, session, redirect, url_for, flash
from flask_cors import CORS
from db import get_db
import pandas as pd
import os
import tempfile
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from werkzeug.security import check_password_hash, generate_password_hash
from crms.routes import crms_bp, db, mail
from fsrm.app import fsrm_bp
from dotenv import load_dotenv
from pathlib import Path

# ----------------------------
# 🔁 Load .env files
# ----------------------------

# 1. Load Project-level .env (Project-CSM/.env)
load_dotenv(dotenv_path=Path(__file__).resolve().parent / ".env")

# 2. Load CRMS-level .env (Project-CSM/crms/.env) — will override if same key exists
load_dotenv(dotenv_path=Path(__file__).resolve().parent / "crms" / ".env")



app = Flask(__name__)
CORS(app)

app.secret_key = os.getenv("SECRET_KEY", "your_fallback_secret_key")

# Get from .env (don't hardcode)
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv(
    "SQLALCHEMY_DATABASE_URI",
    "mysql+pymysql://root:root123@localhost:3307/crms_db"
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# ----------------------------
# ✅ Mail Configuration
# ----------------------------
app.config['MAIL_SERVER'] = os.getenv("MAIL_SERVER", "smtp.gmail.com")
app.config['MAIL_PORT'] = int(os.getenv("MAIL_PORT", 587))
app.config['MAIL_USE_TLS'] = os.getenv("MAIL_USE_TLS", "true").lower() == "true"
app.config['MAIL_USERNAME'] = os.getenv("MAIL_USERNAME")
app.config['MAIL_PASSWORD'] = os.getenv("MAIL_PASSWORD")
app.config['MAIL_DEFAULT_SENDER'] = os.getenv("MAIL_DEFAULT_SENDER")

# Debug (Optional)
# print("MAIL_USERNAME:", app.config['MAIL_USERNAME'])
# print("MAIL_DEFAULT_SENDER:", app.config['MAIL_DEFAULT_SENDER'])

db.init_app(app)
mail.init_app(app)

app.register_blueprint(crms_bp, url_prefix='/crms')
app.register_blueprint(fsrm_bp, url_prefix='/fsrm')

# Dummy credentials (you can later link this to your DB)
USERNAME = 'admin'
PASSWORD = 'password123'

UPLOAD_FOLDER = tempfile.gettempdir()
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
excel_data = {}  # Cache for uploaded Excel files

# ------------------ Helpers ---------------------------
def get_projects_for_user(user_id):
    cur = get_db().cursor()
    cur.execute("SELECT id, name FROM projects WHERE assigned_user_id = %s", (user_id,))
    result = cur.fetchall()
    cur.close()
    return result

def get_all_projects():
    cur = get_db().cursor()
    cur.execute("SELECT id, name FROM projects ORDER BY id")
    rows = cur.fetchall()           # list[(id, name)]
    cur.close()
    return rows

def next_default_project_name():
    cur = get_db().cursor()
    cur.execute("SELECT COUNT(*) FROM projects")
    n = cur.fetchone()[0] + 1
    cur.close()
    return f"Project {n}"

def get_admin_projects():                       # all projects
    cur = get_db().cursor()
    cur.execute("SELECT id, name FROM projects ORDER BY id")
    rows = cur.fetchall()
    cur.close()
    return rows

def get_assigned_projects(user_id):             # only mine
    cur = get_db().cursor()
    cur.execute("SELECT id, name FROM projects WHERE owner_id = %s", (user_id,))
    rows = cur.fetchall()
    cur.close()
    return rows

@app.route('/')
def home():
    if 'username' in session:
        return redirect(url_for('index'))  # or your file upload route
    return redirect(url_for('login'))

# other imports...

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password']
        role = request.form['role'] or 'user'

        if not username or not password:
            flash('Username and password are required', 'danger')
            return redirect(url_for('register'))

        cur = get_db().cursor()
        cur.execute("SELECT id FROM users WHERE username = %s", (username,))
        if cur.fetchone():
            flash('Username already exists', 'warning')
            cur.close()
            return redirect(url_for('register'))

        hashed_pw = generate_password_hash(password)
        cur.execute("INSERT INTO users (username, password, role) VALUES (%s, %s, %s)",
                    (username, hashed_pw, role))
        get_db().commit()
        cur.close()

        flash('Registration successful! Please log in.', 'success')
        return redirect(url_for('login'))

    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password_input = request.form['password']

        cur = get_db().cursor()
        cur.execute("SELECT id, password, role FROM users WHERE username = %s", (username,))
        user = cur.fetchone()
        cur.close()

        if user and check_password_hash(user['password'], password_input):
            session['user_id'] = user['id']
            session['username'] = username
            session['role'] = user['role']
            flash("Login successful!", "success")

            if user['role'] == 'admin':
                return redirect(url_for('dashboard'))
            else:
                return redirect(url_for('user_dashboard'))
        else:
            flash('Invalid username or password', 'danger')

    return render_template('login.html')

@app.route('/user_dashboard')
def user_dashboard():
    if 'username' not in session:
        return redirect(url_for('login'))

    user_id = session.get('user_id')
    print("Logged-in User ID:", session.get('user_id'))

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, name FROM projects WHERE owner_id = %s", (user_id,))
    projects = cur.fetchall()
    cur.close()

    projects = get_projects_for_user(user_id)  # Example function
    return render_template("user_dashboard.html", username=session['username'], projects=projects)

@app.route('/dashboard')
def dashboard():
    if 'username' not in session:
        return redirect(url_for('login'))

    # non-admins should never see the admin dashboard
    if session['role'] != 'admin':
        return redirect(url_for('user_dashboard'))

    projects = get_all_projects()      # every project
    return render_template(
        'dashboard.html',
        username=session['username'],
        projects=projects,             # full list
        role='admin'
    )

@app.route('/assign_project', methods=['GET', 'POST'])
def assign_project():
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = get_db()
    cur = conn.cursor()

    if request.method == 'POST':
        try:
            project_id = int(request.form.get('project_id'))
            user_id = int(request.form.get('user_id'))
            cur.execute("UPDATE projects SET owner_id = %s WHERE id = %s", (user_id, project_id))
            conn.commit()
            flash("✅ Project assigned successfully!")
        except Exception as e:
            flash(f"❌ Error: {e}")

    cur.execute("SELECT id, name FROM projects")
    projects = cur.fetchall()

    cur.execute("SELECT id, username FROM users")
    users = cur.fetchall()

    cur.close()
    return render_template("assign_project.html", projects=projects, users=users)


@app.route('/add_project', methods=['GET', 'POST'])
def add_project():
    if 'username' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        raw = request.form.get('project_name', '').strip()
        name = raw or next_default_project_name()

        cur = get_db().cursor()
        cur.execute("SELECT 1 FROM projects WHERE name=%s", (name,))
        if cur.fetchone():
            flash('Project name already exists', 'danger')
            cur.close()
            return redirect(url_for('dashboard'))

        cur.execute(
            "INSERT INTO projects (name, created_by) VALUES (%s, %s)",
            (name, session['user_id'])
        )
        get_db().commit()
        cur.close()

        flash(f'Project “{name}” added!', 'success')
        return redirect(url_for('dashboard'))

    return render_template('add_project.html')

@app.route('/work_products/<int:project_id>')
def work_products(project_id):
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('work_products.html', project_id=project_id)

UPLOAD_FOLDER = tempfile.gettempdir()
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
excel_data = {}

@app.route('/logout')
def logout():
    session.clear()
    flash("Logged out successfully!", "info")
    return redirect(url_for('login'))

@app.route('/safe_ops/<int:project_id>')
def show_products(project_id):
    return render_template('safe_ops_index.html', project_id=project_id)


@app.route('/index')  # or your actual upload page
def index():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('index.html')  # Your choose file page


@app.route("/upload", methods=["POST"])
def upload():
    file = request.files["file"]
    if file and file.filename.endswith((".xlsx", ".xls")):
        filepath = os.path.join(app.config["UPLOAD_FOLDER"], file.filename)
        file.save(filepath)
        excel_data[file.filename] = {"path": filepath}
        xl = pd.ExcelFile(filepath)
        return jsonify({"message": "Uploaded", "filename": file.filename, "sheets": xl.sheet_names})
    return jsonify({"error": "Invalid file format"}), 400

@app.route("/edit", methods=["GET"])
def edit():
    filename = request.args.get("filename")
    sheet = request.args.get("sheet")
    file_info = excel_data.get(filename)

    if not file_info or not os.path.exists(file_info["path"]):
        return jsonify({"error": "File not found"}), 404

    df = pd.read_excel(file_info["path"], sheet_name=sheet, dtype=str).fillna("")
    wb = load_workbook(file_info["path"], data_only=True)
    ws = wb[sheet]

    dropdowns = {}
    if ws.data_validations:
        for dv in ws.data_validations.dataValidation:
            if dv.type == "list" and dv.formula1:
                options = []
                if dv.formula1.startswith("="):
                    try:
                        ref = dv.formula1.strip("=").replace('$', '')
                        if "!" in ref:
                            sheetname, ref = ref.split("!")
                            target_ws = wb[sheetname]
                        else:
                            target_ws = ws

                        min_col, min_row, max_col, max_row = range_boundaries(ref)
                        for row in target_ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                            for cell in row:
                                if cell.value:
                                    options.append(str(cell.value))
                    except Exception:
                        options = []
                else:
                    options = dv.formula1.strip('"').split(",")

                for cell_range in dv.sqref.ranges:
                    min_col, min_row, max_col, max_row = range_boundaries(str(cell_range))
                    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                        for cell in row:
                            dropdowns[cell.coordinate] = options

    return jsonify({
        "columns": df.columns.tolist(),
        "data": df.to_dict(orient="records"),
        "dropdowns": dropdowns
    })

@app.route("/save", methods=["POST"])
def save():
    data = request.json
    filename = data.get("filename")
    sheet = data.get("sheet")
    edited_data = data.get("data")

    if filename not in excel_data:
        return jsonify({"error": "File not found"}), 404

    filepath = excel_data[filename]["path"]
    df = pd.DataFrame(edited_data)

    wb = load_workbook(filepath)
    ws = wb[sheet]

    # Clear existing rows (except header)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    for r, row_data in enumerate(df.values, start=2):
        for c, value in enumerate(row_data, start=1):
            ws.cell(row=r, column=c, value=value)

    wb.save(filepath)
    return jsonify({"message": "Saved successfully"})

@app.route("/download", methods=["GET"])
def download():
    filename = request.args.get("filename")
    custom_name = request.args.get("custom_name", "Edited_File.xlsx")
    if filename in excel_data:
        filepath = excel_data[filename]["path"]
        return send_file(filepath, as_attachment=True, download_name=custom_name)
    return jsonify({"error": "File not found"}), 404

if __name__ == "__main__":
    app.run(debug=True)